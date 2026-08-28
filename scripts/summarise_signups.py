#!/usr/bin/env python3
"""Count volunteer sign-ups per role from the form export, and nothing else.

The volunteer form collects names, email addresses, LinkedIn URLs, affiliations
and free-text answers. None of that belongs in this repository: nobody has been
contacted yet, and the repository is going to be public. What the organizing
team actually needs from the export is which roles people picked and how many
picked each, so that is all this reads.

    python scripts/summarise_signups.py ~/Downloads/"SciPy India 2026 Conference Planning.xlsx"
    python scripts/summarise_signups.py <file> --write

`--write` updates the `signups:` numbers in config/workgroups.yaml in place. It
touches nothing else in that file, so descriptions and aliases you have edited
by hand survive.

The spreadsheet is never copied into the repository, and this script prints no
column other than the roles one. Roles it cannot match against the registry are
reported by name so you can add an alias, since a silently dropped role would
look like nobody signed up for it.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "src"))

from scipy_india_kg.workgroups import load_registry  # noqa: E402

# The form question whose answers we read. Matched loosely because form wording
# gets edited; everything else in the sheet is deliberately ignored.
ROLE_COLUMN = re.compile(r"volunteer roles.*interested", re.IGNORECASE)
SHEET_NAME = re.compile(r"volunteer", re.IGNORECASE)


def read_roles(path: Path) -> list[list[str]]:
    """One list of role names per respondent. Reads a single column."""
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets = [name for name in workbook.sheetnames if SHEET_NAME.search(name)]
    if not sheets:
        raise SystemExit(f"No volunteer sheet in {path.name}. Sheets: {workbook.sheetnames}")
    sheet = workbook[sheets[0]]

    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    matches = [i for i, cell in enumerate(header) if cell and ROLE_COLUMN.search(str(cell))]
    if not matches:
        raise SystemExit(
            f"No column matching {ROLE_COLUMN.pattern!r} in {sheets[0]!r}. "
            f"Headers: {[str(c)[:40] for c in header if c]}"
        )
    column = matches[0]

    people = []
    for row in rows:
        cell = row[column] if column < len(row) else None
        if not cell:
            continue
        parts = [part.strip(" .;") for part in re.split(r"[\n,]+", str(cell))]
        people.append([part for part in parts if part])
    return people


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, help="the form export, kept outside this repo")
    parser.add_argument(
        "--write", action="store_true", help="update signups: in config/workgroups.yaml"
    )
    args = parser.parse_args()

    if not args.workbook.is_file():
        raise SystemExit(f"No such file: {args.workbook}")

    registry = load_registry()
    people = read_roles(args.workbook)

    counts: Counter[str] = Counter()
    unmatched: Counter[str] = Counter()
    for roles in people:
        for role in roles:
            slug = registry.resolve(role)
            if slug:
                counts[slug] += 1
            else:
                unmatched[role] += 1

    multi = sum(1 for roles in people if len(roles) > 1)
    print(f"{len(people)} respondents, {sum(len(r) for r in people)} role selections")
    print(f"{multi} picked more than one role\n")
    for workgroup in registry:
        print(f"  {counts.get(workgroup.slug, 0):3}  {workgroup.name}")

    if unmatched:
        print("\nNot in config/workgroups.yaml. Add a slug or an alias for each:")
        for role, count in unmatched.most_common():
            print(f"  {count:3}  {role}")

    if not args.write:
        print("\nRe-run with --write to update config/workgroups.yaml.")
        return 1 if unmatched else 0

    config = REPO_ROOT / "config" / "workgroups.yaml"
    lines = config.read_text(encoding="utf-8").splitlines()
    slug = None
    for index, line in enumerate(lines):
        match = re.match(r"^\s*-\s*slug:\s*(\S+)", line)
        if match:
            slug = match.group(1)
        elif slug and re.match(r"^\s*signups:", line):
            indent = line[: len(line) - len(line.lstrip())]
            lines[index] = f"{indent}signups: {counts.get(slug, 0)}"
    config.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"\nUpdated {config.relative_to(REPO_ROOT)}.")
    return 1 if unmatched else 0


if __name__ == "__main__":
    sys.exit(main())
